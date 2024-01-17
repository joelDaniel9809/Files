
from PyQt5 import QtCore, QtGui, QtWidgets
from AddService import Ui_AddService
import subprocess
import winreg
from openpyxl import Workbook
import openpyxl
from os.path import exists
 #Tabla No editable
    self.tableWidget.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
    
    
    #Tamaño de las columnas de la table
    header = self.tableWidget.horizontalHeader()
    header.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
    header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
    
    if exists("Services.xlsx"):
        self.Log = openpyxl.load_workbook("Services.xlsx")    
    else:
        self.Log = Workbook()
        # self.Services_Sheet = Log.create_sheet("Services_Sheet")
        self.Log["Sheet"]["A1"] = 0
        self.Log["Sheet"]["B1"] = 0
        self.Log["Sheet"].column_dimensions["A"].width = 40
        self.Log.save("Services.xlsx") 
    
    self.actionAdd_Service.triggered.connect(self.AddSWindow)
    self.actionCheck_Database.triggered.connect(self.OpenExcel)

   
def AddSWindow(self):
    self.AddService = QtWidgets.QDialog()
    self.ui = Ui_AddService()
    self.ui.setupUi(self.AddService)
    self.AddService.show()
    
def OpenExcel(self):
    handle = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
        r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\excel.exe")
    num_values = winreg.QueryInfoKey(handle)[1]
    path =winreg.EnumValue(handle, 0)[1]
    # subprocess.Popen(["C:\\Program Files (x86)\\Microsoft Office\\root\\Office16\\EXCEL.EXE", "/t", "H:/Clinic/Services.xlsx"])
    subprocess.Popen([path, "/t", "F:/Clinic/Services.xlsx"])
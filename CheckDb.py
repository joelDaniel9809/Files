# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'CheckDb.ui'
#
# Created by: PyQt5 UI code generator 5.9.2
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
from os.path import exists
from openpyxl import Workbook,load_workbook

class Ui_Dialog(object):
    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        width = 433
        height = 357
        Dialog.setFixedSize(width, height)
        self.buttonBox = QtWidgets.QDialogButtonBox(Dialog)
        self.buttonBox.setGeometry(QtCore.QRect(80, 300, 341, 32))
        self.buttonBox.setOrientation(QtCore.Qt.Horizontal)
        self.buttonBox.setStandardButtons(QtWidgets.QDialogButtonBox.Cancel|QtWidgets.QDialogButtonBox.Ok)
        self.buttonBox.setObjectName("buttonBox")
        self.tableWidget = QtWidgets.QTableWidget(Dialog)
        self.tableWidget.setGeometry(QtCore.QRect(10, 10, 410, 281))
        self.tableWidget.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.tableWidget.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContentsOnFirstShow)
        self.tableWidget.setAlternatingRowColors(True)
        Dialog.setWindowIcon(QtGui.QIcon('Logo.ico'))
        self.tableWidget.setWordWrap(False)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(2)
        
        self.label_3 = QtWidgets.QLabel(Dialog)
        self.label_3.setGeometry(QtCore.QRect(20, 300, 211, 51))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_3.setFont(font)
        self.label_3.setTextFormat(QtCore.Qt.AutoText)
        self.label_3.setScaledContents(False)
        self.label_3.setWordWrap(True)
        self.label_3.setObjectName("label_3")
        self.checkBox = QtWidgets.QCheckBox(Dialog)
        self.checkBox.setGeometry(QtCore.QRect(270, 330, 141, 17))
        self.checkBox.setObjectName("checkBox")

        
    
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignVCenter)
        self.tableWidget.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(1, item)
        self.tableWidget.horizontalHeader().setStretchLastSection(False)

        self.retranslateUi(Dialog)
        self.buttonBox.accepted.connect(Dialog.accept)
        self.buttonBox.rejected.connect(Dialog.reject)
        QtCore.QMetaObject.connectSlotsByName(Dialog)
        
        ##Style Sheets##
        self.buttonBox.setStyleSheet("background-color:#232326;")
        Dialog.setStyleSheet("QWidget{\n"
          "background-color: rgb(0, 3, 18);\n"
          "color:rgb(255, 255, 255);\n"
          "}")
        self.tableWidget.setStyleSheet("QTableView::item:selected { color:white; background:#000000; font-weight:900; }\n"
            "QTableCornerButton::section { background-color:#232326; }\n"
            "QHeaderView::section { color:white; background-color:#232326; }\n"
            "QTableView::item{\n"
            "    background-color: #232326;\n"
            "}\n"
            "        ")
        
        
        
    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Check Database"))
  
        item = self.tableWidget.horizontalHeaderItem(0)
        item.setText(_translate("Dialog", "Service"))
        item = self.tableWidget.horizontalHeaderItem(1)
        item.setText(_translate("Dialog", "   Price   "))
        self.label_3.setText(_translate("CheckDb", "Tip: Double Click on any row of the table to delete it."))
        self.checkBox.setText(_translate("CheckDb", "Edit Values"))

        
        #Tabla No editable
        self.tableWidget.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
        self.tableWidget.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        
        if exists("Services.xlsx"):
            self.Log = load_workbook("Services.xlsx")    
        else:
            self.Log = Workbook()
            self.Log["Sheet"]["A1"] = ""
            self.Log["Sheet"]["B1"] = ""
            self.Log["Sheet"].column_dimensions["A"].width = 40
            self.Log.save("Services.xlsx") 
        
        self.ws = self.Log["Sheet"]
        self.Names = list(self.ws.columns)[0]
        self.Prices= list(self.ws.columns)[1]
        
        
        cont = 0;
        for i in self.Names:
            
            self.tableWidget.setRowCount(cont+1)
            item = QtWidgets.QTableWidgetItem()
            self.tableWidget.setVerticalHeaderItem(cont, item)
            item = self.tableWidget.verticalHeaderItem(cont)
            item.setText(_translate("Dialog", str(cont+1)))
            
            self.tableWidget.setItem(cont, 0, QtWidgets.QTableWidgetItem(i.value))
            cont = cont+1
        cont = 0
        for i in self.Prices:
            item= QtWidgets.QTableWidgetItem(i.value)
            item.setTextAlignment(QtCore.Qt.AlignHCenter|QtCore.Qt.AlignVCenter)
            self.tableWidget.setItem(cont, 1, item)
            cont= cont+1
        
        
        #Tamaño de las columnas de la tabla
        header = self.tableWidget.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
        
        
        x= self.tableWidget.doubleClicked.connect(self.DeleteNames)
        self.buttonBox.accepted.connect(self.UpdateDb)
        self.checkBox.stateChanged.connect(self.ChangeTableProperties)
        x= self.tableWidget.cellChanged.connect(self.autoCapital)
        
        self.Changed = False
        
    def autoCapital(self , x):
        txt = self.tableWidget.item(x,0).text()
        if (txt.isupper() == True):
            return
        upper = txt.upper()
        self.tableWidget.setItem(x, 0, QtWidgets.QTableWidgetItem(upper))
        
    def ChangeTableProperties(self, state):
        if (QtCore.Qt.Checked == state):
            self.Changed = True
            self.tableWidget.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectItems)
            self.tableWidget.setEditTriggers(QtWidgets.QTableWidget.AllEditTriggers)
        else:
            self.tableWidget.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
            self.tableWidget.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        
    def UpdateDb(self):
        if (self.Changed == True):
            self.ws.delete_rows(1, self.ws.max_row+1)
            self.Log["Sheet"]["A1"] = ""
            self.Log["Sheet"]["B1"] = ""
            for i in range(self.tableWidget.rowCount()):
                self.ws[i+1][0].value = self.tableWidget.item(i,0).text()
                self.ws[i+1][1].value = self.tableWidget.item(i,1).text()
            self.Log.save("Services.xlsx")
            
        
    def DeleteNames(self, x):
        self.Changed = True
        row = x.row()
        self.tableWidget.removeRow(row)
        
        
        
        


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Dialog = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(Dialog)
    Dialog.show()
    sys.exit(app.exec_())


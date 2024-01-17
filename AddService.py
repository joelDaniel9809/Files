# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'AddService.ui'
#
# Created by: PyQt5 UI code generator 5.9.2
#
# WARNING! All changes made in this file will be lost!


from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import Workbook,load_workbook
from os.path import exists


class Ui_AddService(object):
    def setupUi(self, AddService):
        AddService.setObjectName("AddService")
        width = 398
        height = 113
        AddService.setFixedSize(width, height)
        AddService.setWindowIcon(QtGui.QIcon('Logo.ico'))
        self.gridLayoutWidget_2 = QtWidgets.QWidget(AddService)
        self.gridLayoutWidget_2.setGeometry(QtCore.QRect(10, 10, 381, 91))
        self.gridLayoutWidget_2.setObjectName("gridLayoutWidget_2")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.gridLayoutWidget_2)
        self.gridLayout_2.setContentsMargins(0, 0, 0, 0)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.label_6 = QtWidgets.QLabel(self.gridLayoutWidget_2)
        self.label_6.setEnabled(True)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Maximum)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_6.sizePolicy().hasHeightForWidth())
        self.label_6.setSizePolicy(sizePolicy)
        self.label_6.setMaximumSize(QtCore.QSize(16777215, 20))
        self.label_6.setObjectName("label_6")
        self.gridLayout_2.addWidget(self.label_6, 0, 0, 1, 1)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.pushButton_2 = QtWidgets.QPushButton(self.gridLayoutWidget_2)
        self.pushButton_2.setEnabled(True)
        self.pushButton_2.setObjectName("pushButton_2")
        self.horizontalLayout.addWidget(self.pushButton_2)
        self.label = QtWidgets.QLabel(self.gridLayoutWidget_2)
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label.setObjectName("label")
        self.horizontalLayout.addWidget(self.label)
        self.horizontalLayout.setStretch(0, 2)
        self.horizontalLayout.setStretch(1, 3)
        self.horizontalLayout_2.addLayout(self.horizontalLayout)
        self.gridLayout_2.addLayout(self.horizontalLayout_2, 2, 0, 1, 1)
        self.label_5 = QtWidgets.QLabel(self.gridLayoutWidget_2)
        self.label_5.setEnabled(True)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Maximum)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_5.sizePolicy().hasHeightForWidth())
        self.label_5.setSizePolicy(sizePolicy)
        self.label_5.setMaximumSize(QtCore.QSize(16777215, 20))
        self.label_5.setObjectName("label_5")
        self.gridLayout_2.addWidget(self.label_5, 0, 1, 1, 1)
        self.lineEdit = QtWidgets.QLineEdit(self.gridLayoutWidget_2)
        self.lineEdit.setObjectName("lineEdit")
        self.gridLayout_2.addWidget(self.lineEdit, 1, 0, 1, 1)
        self.lineEdit_2 = QtWidgets.QLineEdit(self.gridLayoutWidget_2)
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.gridLayout_2.addWidget(self.lineEdit_2, 1, 1, 1, 1)
        self.gridLayout_2.setRowMinimumHeight(0, 2)
        self.gridLayout_2.setRowMinimumHeight(1, 1)
        self.gridLayout_2.setRowMinimumHeight(2, 1)
        self.gridLayout_2.setColumnStretch(0, 3)
        self.gridLayout_2.setColumnStretch(1, 1)

        self.retranslateUi(AddService)
        QtCore.QMetaObject.connectSlotsByName(AddService)
        self.pushButton_2.setStyleSheet("background-color:#232326;")
        AddService.setStyleSheet("QWidget{\n"
          "background-color: rgb(0, 3, 18);\n"
          "color:rgb(255, 255, 255);\n"
          "}")
        self.lineEdit.setStyleSheet("background-color:#232326;\n"
"    color:rgb(255, 255, 255)")
        self.lineEdit_2.setStyleSheet("background-color:#232326;\n"
"    color:rgb(255, 255, 255)")        
        

    def retranslateUi(self, AddService):
        _translate = QtCore.QCoreApplication.translate
        AddService.setWindowTitle(_translate("AddService", "Add Service to Database"))
        self.label_6.setText(_translate("AddService", "Service Name"))
        self.pushButton_2.setText(_translate("AddService", "Add"))
        # self.label.setText(_translate("AddService", "Status"))
        self.label_5.setText(_translate("AddService", "Price(USD)"))
    
########Excel Section############
        if exists("Services.xlsx"):
            self.Log = load_workbook("Services.xlsx")    
        else:
            self.Log = Workbook()
            # self.Services_Sheet = Log.create_sheet("Services_Sheet")
            self.Log["Sheet"]["A1"] = ""
            self.Log["Sheet"]["B1"] = ""
            self.Log["Sheet"].column_dimensions["A"].width = 40
            self.Log.save("Services.xlsx") 
#################################
        self.pushButton_2.clicked.connect(self.CheckServiceAdd)
        self.lineEdit.textChanged.connect(self.auto_capital)
        self.lineEdit_2.returnPressed.connect(self.lineEdit.setFocus)
    
    
    def ClearLabel (self):
        self.label.setText("")
        
    def auto_capital(self , txt):
        upper = txt.upper()
        self.lineEdit.setText(upper)
    
    def CheckServiceAdd(self):
        
        if (str(self.lineEdit.text()) and str(self.lineEdit_2.text()) ):
            self.ws = self.Log["Sheet"]
            self.Names = list(self.ws.columns)[0]
            self.Num = 1
            for i in self.Names:
                if (i.value == self.lineEdit.text()):
                    self.label.setText("Service already in Database")
                    return
                if (i.value):
                    self.Num+=1
                    continue
                break
            self.ws[self.Num][0].value = self.lineEdit.text()
            self.ws[self.Num][1].value = self.lineEdit_2.text()
            self.lineEdit.clear()
            self.lineEdit_2.clear()
           
            timer = QtCore.QTimer()
            timer.singleShot(2000, self.ClearLabel)
            self.label.setText("Service Added")
            self.Log.save("Services.xlsx") 
            
        elif (str(self.lineEdit.text())=="" or str(self.lineEdit_2.text())==""):
            self.label.setText("Check Info")
        


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    AddService = QtWidgets.QDialog()
    ui = Ui_AddService()
    ui.setupUi(AddService)
    AddService.show()
    sys.exit(app.exec_())


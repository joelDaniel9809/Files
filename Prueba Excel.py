from openpyxl import Workbook
import openpyxl
from os.path import exists
import re

# if exists("Services.xlsx"):
#     Log = openpyxl.load_workbook("Services.xlsx")        
# else:
#     Log = Workbook() 
#     Log.save("Services.xlsx") 
    
# ws = Log["Sheet"]
# ws[1][0].value = 1
# ws[2][0].value = 1
# Log.save("Services.xlsx") 

import winreg


handle = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
    r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\excel.exe")
num_values = winreg.QueryInfoKey(handle)[1]
path =winreg.EnumValue(handle, 0)[1]
slash = '\'
dslash = '\\'
path = path.replace(slash , dslash)
# print(path)
# for i in range(num_values):
#     print(winreg.EnumValue(handle, i))
#     print (i)
# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'Main.ui'
#
# Created by: PyQt5 UI code generator 5.9.2
#
# WARNING! All changes made in this file will be lost!
from PyQt5 import QtCore, QtGui, QtWidgets
from AddService import Ui_AddService
from openpyxl import Workbook , load_workbook
from os.path import exists
from CheckDb import Ui_Dialog
from datetime import date
from os import startfile,path
from tempfile import mktemp
from PIL.Image import new
from PIL.Image import open as op
from PIL.ImageDraw import Draw
from PIL.ImageFont import truetype


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        width = 446
        height = 341
        MainWindow.setFixedSize(width, height)
        MainWindow.setAutoFillBackground(True)
        MainWindow.setWindowIcon(QtGui.QIcon('Logo.ico'))
        
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.layoutWidget = QtWidgets.QWidget(self.centralwidget)
        self.layoutWidget.setGeometry(QtCore.QRect(10, 10, 421, 251))
        self.layoutWidget.setObjectName("layoutWidget")
        self.gridLayout = QtWidgets.QGridLayout(self.layoutWidget)
        self.gridLayout.setContentsMargins(0, 0, 0, 0)
        self.gridLayout.setObjectName("gridLayout")
        self.label = QtWidgets.QLabel(self.layoutWidget)
        self.label.setMaximumSize(QtCore.QSize(16777215, 22))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(12)
        self.label.setFont(font)
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label.setObjectName("label")
        self.gridLayout.addWidget(self.label, 0, 0, 1, 1)
        self.label_2 = QtWidgets.QLabel(self.layoutWidget)
        self.label_2.setMaximumSize(QtCore.QSize(16777215, 22))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(12)
        self.label_2.setFont(font)
        self.label_2.setAlignment(QtCore.Qt.AlignCenter)
        self.label_2.setObjectName("label_2")
        self.gridLayout.addWidget(self.label_2, 0, 1, 1, 1)
        
        self.tableWidget = QtWidgets.QTableWidget(self.layoutWidget)
        self.tableWidget.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.tableWidget.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOn)
        self.tableWidget.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContentsOnFirstShow)
        self.tableWidget.setAlternatingRowColors(True)
        self.tableWidget.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.tableWidget.setWordWrap(False)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(2)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignVCenter)
        self.tableWidget.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(1, item)
        self.tableWidget.horizontalHeader().setStretchLastSection(False)
        self.gridLayout.addWidget(self.tableWidget, 2, 0, 1, 2)
        self.lineEdit = QtWidgets.QLineEdit(self.layoutWidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineEdit.sizePolicy().hasHeightForWidth())
        self.lineEdit.setSizePolicy(sizePolicy)
        self.lineEdit.setMaximumSize(QtCore.QSize(200, 16777215))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.lineEdit.setFont(font)
        self.lineEdit.setObjectName("lineEdit")
        self.gridLayout.addWidget(self.lineEdit, 1, 1, 1, 1)
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(237, 270, 191, 27))
        self.pushButton.setObjectName("pushButton")
        self.lineEdit_2 = QtWidgets.QLineEdit(self.layoutWidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineEdit_2.sizePolicy().hasHeightForWidth())
        self.lineEdit_2.setSizePolicy(sizePolicy)
        self.lineEdit_2.setMaximumSize(QtCore.QSize(200, 16777215))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.lineEdit_2.setFont(font)
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.gridLayout.addWidget(self.lineEdit_2, 1, 0, 1, 1)
        self.gridLayout.setColumnStretch(0, 3)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 446, 21))
        self.menubar.setObjectName("menubar")
        self.menuServicios = QtWidgets.QMenu(self.menubar)
        self.menuServicios.setObjectName("menuServicios")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.actionAdd_Service = QtWidgets.QAction(MainWindow)
        self.actionAdd_Service.setObjectName("actionAdd_Service")
        self.actionCheck_Database = QtWidgets.QAction(MainWindow)
        self.actionCheck_Database.setObjectName("actionCheck_Database")
        self.menuServicios.addAction(self.actionAdd_Service)
        self.menuServicios.addAction(self.actionCheck_Database)
        self.menubar.addAction(self.menuServicios.menuAction())
        
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(20, 261, 211, 51))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_3.setFont(font)
        self.label_3.setTextFormat(QtCore.Qt.AutoText)
        self.label_3.setScaledContents(False)
        self.label_3.setWordWrap(True)
        self.label_3.setObjectName("label_3")

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
        MainWindow.setTabOrder(self.lineEdit_2, self.lineEdit)
        MainWindow.setTabOrder(self.lineEdit, self.pushButton)
        MainWindow.setTabOrder(self.pushButton, self.tableWidget)
#####StyleSheets####
        MainWindow.setStyleSheet("QWidget{\n"
         "background-color: rgb(0, 3, 18);\n"
         "color:rgb(255, 255, 255);\n"
         "}")
        self.tableWidget.setStyleSheet("QTableView::item:selected { color:white; background:#000000; font-weight:900; }\n"
"QTableCornerButton::section { background-color:#232326; }\n"
"QHeaderView::section { color:white; background-color:#232326; }\n"
"QTableView::item{\n"
"    background-color: #232326;\n"
"}\n"
"        ")
        self.lineEdit.setStyleSheet("background-color:#232326;\n"
"    color:rgb(255, 255, 255)")
        self.lineEdit_2.setStyleSheet("background-color:#232326;\n"
"    color:rgb(255, 255, 255)")
        self.pushButton.setStyleSheet("background-color: #232326")        
        
    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Clinic Services"))
        self.label.setText(_translate("MainWindow", "Client Name"))
        self.label_2.setText(_translate("MainWindow", "Add Service"))
        item = self.tableWidget.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Service"))
        item = self.tableWidget.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "    Price    "))
        self.menuServicios.setTitle(_translate("MainWindow", "Services"))
        self.pushButton.setText(_translate("MainWindow", "Print"))
        self.actionAdd_Service.setText(_translate("MainWindow", "Add Service to Database"))
        self.actionCheck_Database.setText(_translate("MainWindow", "Check Database"))
        self.label_3.setText(_translate("MainWindow", "Tip: Double Click on any row of the table to delete it."))
    
    
        #Tabla No editable
        self.tableWidget.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
        
        
        #Tamaño de las columnas de la table
        header = self.tableWidget.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
        
        self.Completer()
        
        self.actionAdd_Service.triggered.connect(self.AddSWindow)
        self.actionCheck_Database.triggered.connect(self.OpenDb)
        
        
        self.lineEdit.returnPressed.connect(self.AddToTable)
        self.lineEdit.textChanged.connect(self.auto_capital)
        self.tableWidget.cellChanged.connect(self.Timer)
        self.tableWidget.cellChanged.connect(self.SetFullPrice)
        x= self.tableWidget.doubleClicked.connect(self.DeleteNames)
        self.pushButton.clicked.connect(self.CreateTicket)
        
        
        
   
        
   #####Definitions###
        self.CountServices = 0
    def SetFullPrice(self):
        self.FullPrice = 0
        if(not self.CountServices):
            self.label_3.setText("Full Price:0 USD")
            return
        
        for i in range(self.CountServices):
            Price = self.tableWidget.item(i,1)
            self.FullPrice+= float(Price.text())
            self.label_3.setText("Full Price: %.2f USD"%self.FullPrice)
            
            
        
    
        
    def Timer(self):
        timer = QtCore.QTimer()
        timer.singleShot(0, self.ClearLEdit)
    def ClearLEdit(self):
        self.lineEdit.clear()

    def CreateTicket (self):
        if (self.lineEdit_2.text()== ""):
            self.label_3.setText("Check Client Name")
        elif(not self.CountServices):
            self.label_3.setText("No Services Listed")
        else:
            f= open("ticket.txt","w+")
            f.write("Benech Family Clinic\n")
            f.write("Client:%s \n"% (self.lineEdit_2.text()))
            Date = date.today()
            Date = Date.strftime("%d-%m-%Y")
            f.write("Date: %s \n"% Date )
            f.write("\nSERVICES                           PRICES\n")
        
            for i in range(self.CountServices):
                Service = self.tableWidget.item(i,0)
                Space=35 - len(str(Service.text()))
                Price = self.tableWidget.item(i,1)
                f.write("%s%s%.2f USD\n"% (str(Service.text()),(Space*"-"),float(Price.text())) )
            
            f.write("\nFull Price:------------------------%s USD"% (str(self.FullPrice) ))
            
            self.tableWidget.clearContents()
            
            self.tableWidget.setRowCount(1)
            self.lineEdit_2.clear()
            self.label_3.setText("Full Price:0 USD")
            
            out = new("RGB", (330,200 + 15*self.CountServices), (255, 255, 255)) 
            # get a font
            fnt =truetype("Pillow/Tests/fonts/cour.ttf", 12)
            # get a drawing context
            d = Draw(out)
            f= open("ticket.txt","r")
            im  = op("Logo.jpg")
            out.paste(im , (110,1))
            # draw multiline text
            d.multiline_text((5, 100), f.read(), font=fnt, fill=(0, 0, 0))
            
            out.save("out.png")
            self.CountServices = 0
            
            # out.show()
            startfile("out.png","print")
        
    def AddToTable(self):
        cont = 1
        
        for i in self.Names:
            
            if (i.value == self.lineEdit.text()):
                
                self.tableWidget.setRowCount(self.CountServices +1)
                item = QtWidgets.QTableWidgetItem()
                self.tableWidget.setVerticalHeaderItem(self.CountServices, item)
                
                item = self.tableWidget.verticalHeaderItem(self.CountServices)
                item.setText((str(self.CountServices +1)))
                
                self.tableWidget.setItem(self.CountServices,0,QtWidgets.QTableWidgetItem(self.lineEdit.text()))
                price = self.ws[cont][1].value
                item= QtWidgets.QTableWidgetItem(price)
                item.setTextAlignment(QtCore.Qt.AlignHCenter|QtCore.Qt.AlignVCenter)
                self.CountServices = self.CountServices + 1
                self.tableWidget.setItem(self.CountServices-1,1,item)
                break
            else:
                cont = cont + 1
            self.SetFullPrice()
            
            
            
    def auto_capital(self , txt):
        upper = txt.upper()
        self.lineEdit.setText(upper)
    
    def DeleteNames(self, x):
        row = x.row()
        self.tableWidget.removeRow(row)
        self.CountServices -=1
        self.SetFullPrice()
            
    
    def Completer(self):
        if exists("Services.xlsx"):
            self.Log = load_workbook("Services.xlsx")    
        else:
            self.Log = Workbook()
            # self.Services_Sheet = Log.create_sheet("Services_Sheet")
            self.Log["Sheet"]["A1"] = ""
            self.Log["Sheet"]["B1"] = ""
            self.Log["Sheet"].column_dimensions["A"].width = 40
            self.Log.save("Services.xlsx") 
            
        self.ws = self.Log["Sheet"]
        self.names = []
        self.Names = list(self.ws.columns)[0]
        for i in self.Names:
            self.names.append(i.value) 
        
        completer = QtWidgets.QCompleter(self.names)
        completer.setCompletionRole(QtWidgets.QCompleter.PopupCompletion)
        completer.popup().setStyleSheet("background-color:#232326;\n"
"    color:rgb(255, 255, 255)")
        completer.setFilterMode(QtCore.Qt.MatchContains)
        self.lineEdit.setCompleter(completer)  
    
    def AddSWindow(self):
        self.AddService = QtWidgets.QDialog()
        self.ui = Ui_AddService()
        self.AddService.finished.connect(self.Completer)
        self.ui.setupUi(self.AddService)
        self.AddService.show()
        
    def OpenDb(self):
        self.CheckDB = QtWidgets.QDialog()
        self.CheckDB.finished.connect(self.Completer)
        self.ui = Ui_Dialog()
        self.ui.setupUi(self.CheckDB)
        self.CheckDB.show()

    

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())

